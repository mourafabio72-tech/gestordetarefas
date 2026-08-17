"""Importa a matriz Responsável × Setor das empresas via Excel.

Layout: coluna CNPJ + uma coluna por SETOR (o cabeçalho é o nome do setor).
A célula traz o NOME do responsável daquele setor naquela empresa.
  - célula preenchida -> a empresa ATENDE o setor, com aquele responsável
    (se o nome não casar, atende sem responsável + aviso);
  - célula vazia       -> DESMARCA o setor (não atende, não gera tarefa).
Só mexe nos setores que vierem como coluna na planilha.
"""
import io
import re
import unicodedata
from ..models import Empresa, Usuario, Setor, EmpresaSetorResponsavel


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _so_digitos(s: str) -> str:
    return re.sub(r"\D", "", str(s or ""))


def _linhas(nome_arquivo: str, conteudo: bytes):
    if (nome_arquivo or "").lower().endswith(".xls"):
        import xlrd
        sh = xlrd.open_workbook(file_contents=conteudo).sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    import openpyxl
    ws = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True).worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def gerar_modelo(db) -> bytes:
    import openpyxl
    setores = db.query(Setor).filter(Setor.ativo == True).order_by(Setor.nome).all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Responsáveis"
    ws.append(["CNPJ"] + [s.nome for s in setores])
    ws.append(["12.345.678/0001-90"] + (["Nome do responsável"] if setores else []))
    ws.column_dimensions["A"].width = 22
    for i, _ in enumerate(setores):
        ws.column_dimensions[chr(ord("B") + i)].width = 22
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar(db, nome_arquivo: str, conteudo: bytes) -> dict:
    grade = _linhas(nome_arquivo, conteudo)
    idx = next((i for i, r in enumerate(grade) if any(c not in (None, "") for c in r)), None)
    if idx is None:
        return {"resumo": {"empresas": 0, "marcados": 0, "desmarcados": 0, "erros": 0}, "detalhes": []}

    cabecalho = grade[idx]
    col_cnpj = next((i for i, t in enumerate(cabecalho) if _norm(t) == "cnpj"), None)
    if col_cnpj is None:
        return {"erro": "A planilha precisa de uma coluna 'CNPJ'.",
                "resumo": {"empresas": 0, "marcados": 0, "desmarcados": 0, "erros": 0}, "detalhes": []}

    # coluna -> setor (casado por nome)
    setores = db.query(Setor).all()
    setor_por_nome = {_norm(s.nome): s for s in setores}
    col_setor = {}
    for i, t in enumerate(cabecalho):
        if i == col_cnpj:
            continue
        s = setor_por_nome.get(_norm(t))
        if s:
            col_setor[i] = s

    usuarios = [u for u in db.query(Usuario).filter(Usuario.tipo != "cliente").all()]
    usuario_por_nome = {_norm(u.nome): u for u in usuarios}

    empresas_ok = marcados = desmarcados = erros = 0
    detalhes = []
    for linha in grade[idx + 1:]:
        if not any(c not in (None, "") for c in linha):
            continue
        cnpj = _so_digitos(linha[col_cnpj] if col_cnpj < len(linha) else "")
        emp = next((e for e in db.query(Empresa).filter(Empresa.cnpj != None).all()
                    if _so_digitos(e.cnpj) == cnpj), None) if cnpj else None
        if not emp:
            erros += 1
            detalhes.append({"linha": cnpj or "(sem CNPJ)", "status": "erro", "detalhe": "Empresa não encontrada pelo CNPJ."})
            continue
        for i, setor in col_setor.items():
            valor = str(linha[i]).strip() if (i < len(linha) and linha[i] not in (None, "")) else ""
            existente = (db.query(EmpresaSetorResponsavel)
                         .filter(EmpresaSetorResponsavel.empresa_id == emp.id,
                                 EmpresaSetorResponsavel.setor_id == setor.id).first())
            if valor:
                resp = usuario_por_nome.get(_norm(valor))
                if not resp:
                    detalhes.append({"linha": emp.razao_social, "status": "aviso",
                                     "detalhe": f"{setor.nome}: responsável '{valor}' não encontrado, setor marcado sem responsável."})
                if existente:
                    existente.responsavel_id = resp.id if resp else None
                else:
                    db.add(EmpresaSetorResponsavel(empresa_id=emp.id, setor_id=setor.id,
                                                   responsavel_id=resp.id if resp else None))
                marcados += 1
            else:
                if existente:
                    db.delete(existente)
                    desmarcados += 1
        empresas_ok += 1

    db.commit()
    return {"resumo": {"empresas": empresas_ok, "marcados": marcados, "desmarcados": desmarcados, "erros": erros},
            "detalhes": detalhes}
