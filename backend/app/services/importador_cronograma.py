"""Importação de um CRONOGRAMA de fechamento (ex.: GRABER) como Obrigações
recorrentes. Cada atividade vira uma obrigação vinculada às entidades do grupo,
com prazo em N-ésimo dia útil (convertido da data do arquivo).

Fluxo: analisar(arquivo) -> preview editável -> importar(itens) -> cria empresas
do grupo + setores + obrigações.
"""
import io
import re
import calendar
import unicodedata
from collections import defaultdict, Counter
from datetime import datetime, date
from ..models import Empresa, Setor, Obrigacao

SETORES_PADRAO = ["Contabilidade", "Fiscal", "DP", "Financeiro"]
MESES_TODOS = "1,2,3,4,5,6,7,8,9,10,11,12"


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.lower().strip()


def nth_dia_util(d: date) -> int:
    """Qual dia útil (seg-sex) do mês a data representa."""
    n = 0
    for dia in range(1, d.day + 1):
        if date(d.year, d.month, dia).weekday() < 5:
            n += 1
    return n


def chute_setor(nome: str) -> str:
    """Classificação por palavra-chave. Ordem: Fiscal, Contabilidade, DP,
    Financeiro (o usuário revisa no preview)."""
    n = _norm(nome)
    fiscal = ["icms", "pis", "cofins", " ipi", "iss", "irpj", "csll", "apuracao",
              "impostos retid", "retid", "guia", "sped", "reinf", "difal"]
    contab = ["dre", "balancete", "balanco", "book", "cmv", "imobilizado", "ifrs",
              "deprecia", "amortiza", "intercompany", "contabiliz", "faturamento",
              "rateio", "time sheet", "timesheet", "razao", "provisao", "reversao",
              " bp", "estoque", "selic", "diferido", "deferido"]
    dp = ["folha", "inss", "fgts", "ferias", "13o", "13º", "encargos", "rescis",
          "admissao", "holerite", "ponto"]
    fin = ["fluxo de caixa", "extrato", "baixas", "titulos", "bancari", "banco",
           "aplicacao financeira", "saldo", "contas a pagar", "contas a receber"]
    for setor, chaves in [("Fiscal", fiscal), ("Contabilidade", contab),
                          ("DP", dp), ("Financeiro", fin)]:
        if any(k in n for k in chaves):
            return setor
    return ""  # sem chute -> usuário escolhe


def parse_entidades(cel) -> list:
    """Divide a coluna Empresa em códigos/nomes (ex.: 'FOS, SL' -> ['FOS','SL'])."""
    c = str(cel or "").strip()
    if not c:
        return []
    return sorted({t.strip().upper() for t in re.split(r"[/,;]+", c) if t.strip()})


# ---- Layout único: Descrição da tarefa · Competência · Vencimento · Empresa · Gera multa · Setor ----
FIELD_ALIASES = {
    "nome": ["descricao da tarefa", "descricao das tarefas", "descricao das tarefa",
             "descricao", "descrição", "tarefa", "atividade", "obrigacao", "obrigação"],
    "competencia": ["competencia", "competência"],
    "vencimento": ["vencimento", "prazo", "transicao", "transição", "meta", "data prevista", "data"],
    "empresa": ["empresa", "empresas", "entidade", "entidades"],
    "cnpj": ["cnpj", "cnpj/cpf", "cnpj cpf"],
    "multa": ["gera multa", "multa", "passivel de multa", "passível de multa"],
    "setor": ["setor", "departamento", "area", "área", "responsavel", "responsável"],
}

COMPETENCIAS = {
    "mes anterior": "mes_anterior", "mesmo mes": "mesmo_mes",
    "ano anterior": "ano_anterior", "mes seguinte": "mes_seguinte", "subsequente": "mes_seguinte",
}


def _mapear_colunas(header) -> dict:
    m = {}
    norm = [_norm(c) for c in header]
    for field, aliases in FIELD_ALIASES.items():
        for j, h in enumerate(norm):
            if h and h in aliases and field not in m:
                m[field] = j
    return m


def map_competencia(cel) -> str:
    t = _norm(cel)
    for k, v in COMPETENCIAS.items():
        if k in t:
            return v
    return "mes_anterior"


def parse_vencimento(cel):
    """Devolve (regra_prazo_tipo, regra_prazo_dia, label) de uma data OU de texto
    ('5º dia útil', 'último dia útil')."""
    if isinstance(cel, datetime):
        n = nth_dia_util(cel.date())
        return "dia_util", n, f"{n}º dia útil"
    t = _norm(cel)
    m = re.search(r"(\d+)\s*[ºo°]?\s*dia\s*util", t)
    if m:
        n = int(m.group(1))
        return "dia_util", n, f"{n}º dia útil"
    if "ultimo dia" in t:
        return "ultimo_dia_util", None, "último dia útil"
    if "primeiro dia" in t:
        return "primeiro_dia_util", None, "primeiro dia útil"
    return "ultimo_dia_util", None, "—"


def mapear_setor(cel) -> str:
    """Mapeia o texto da coluna Setor/Responsável para um dos 4 setores padrão."""
    t = _norm(cel)
    if not t:
        return ""
    if "fiscal" in t:
        return "Fiscal"
    if re.search(r"\bdp\b", t) or "pessoal" in t:
        return "DP"
    if "contab" in t or "controlad" in t:
        return "Contabilidade"
    if any(k in t for k in ["pagar", "receber", "tesouraria", "faturamento", "financ", "caixa"]):
        return "Financeiro"
    return ""


def parse_multa(cel) -> bool:
    t = _norm(cel)
    return t in ("sim", "s", "x", "1", "true", "verdadeiro") or "sim" in t


def _ler(conteudo: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    return list(wb.worksheets[0].iter_rows(values_only=True))


def _achar_cabecalho(rows):
    """Cabeçalho = 1ª linha com a coluna de descrição/atividade."""
    for i, r in enumerate(rows):
        m = _mapear_colunas(r)
        if "nome" in m:
            return i, m
    return None, {}


def analisar(conteudo: bytes) -> dict:
    """Lê a planilha (layout único) e devolve as obrigações candidatas, agrupadas por descrição."""
    rows = _ler(conteudo)
    i_cab, c = _achar_cabecalho(rows)
    if i_cab is None:
        return {"erro": "Não encontrei o cabeçalho (precisa da coluna 'Descrição da tarefa').", "itens": []}

    def cel(r, campo):
        j = c.get(campo)
        return r[j] if (j is not None and j < len(r)) else None

    grupos, ordem = {}, []
    for r in rows[i_cab + 1:]:
        nome_cel = cel(r, "nome")
        if not nome_cel or not str(nome_cel).strip():
            continue
        nome = str(nome_cel).strip()
        if nome not in grupos:
            grupos[nome] = {"ents": set(), "cnpjs": set(), "setor": None, "setor_raw": None,
                            "comp": None, "tipo": None, "dia": None, "label": None, "multa": False}
            ordem.append(nome)
        g = grupos[nome]
        g["ents"].update(parse_entidades(cel(r, "empresa")))
        cnpj = re.sub(r"\D", "", str(cel(r, "cnpj") or ""))
        if cnpj:
            g["cnpjs"].add(cnpj)
        sc = cel(r, "setor")
        if g["setor_raw"] is None and sc not in (None, ""):
            g["setor_raw"] = str(sc).strip()
        if g["setor"] is None:
            g["setor"] = mapear_setor(sc) or chute_setor(nome)
        if g["comp"] is None and cel(r, "competencia") is not None:
            g["comp"] = map_competencia(cel(r, "competencia"))
        v = cel(r, "vencimento")
        if g["tipo"] is None and v not in (None, ""):
            g["tipo"], g["dia"], g["label"] = parse_vencimento(v)
        if parse_multa(cel(r, "multa")):
            g["multa"] = True

    itens, entidades = [], set()
    for nome in ordem:
        g = grupos[nome]
        ents = sorted(g["ents"])
        entidades.update(ents)
        itens.append({
            "nome": nome, "setor": g["setor"] or "", "setor_raw": g["setor_raw"] or "",
            "entidades": ents, "cnpjs": sorted(g["cnpjs"]),
            "competencia_ref": g["comp"] or "mes_anterior",
            "regra_prazo_tipo": g["tipo"] or "ultimo_dia_util",
            "regra_prazo_dia": g["dia"], "prazo_label": g["label"] or "—",
            "gera_multa": g["multa"],
        })
    itens.sort(key=lambda x: (x["regra_prazo_dia"] or 99, x["nome"]))
    return {"itens": itens, "entidades": sorted(entidades), "total": len(itens)}


def _get_or_create_empresa(db, codigo: str, grupo: str):
    e = db.query(Empresa).filter(Empresa.razao_social == codigo).first()
    if not e:
        e = Empresa(razao_social=codigo, grupo=grupo, ativo=True)
        db.add(e)
        db.flush()
    elif grupo and not e.grupo:
        e.grupo = grupo
    return e


def _get_or_create_setor(db, nome: str):
    if not nome:
        return None
    s = db.query(Setor).filter(Setor.nome == nome).first()
    if not s:
        s = Setor(nome=nome)
        db.add(s)
        db.flush()
    return s


def importar(db, grupo: str, itens: list, mapa: dict = None) -> dict:
    """Cria setores e obrigações (dedupe por nome; se já existe, apenas acrescenta
    as empresas ao vínculo). Cada código de entidade (FOS/SL/SLS) é resolvido para
    uma empresa: se `mapa[codigo]` aponta uma empresa cadastrada, usa ela; senão,
    cria uma empresa pelo próprio código (fallback)."""
    grupo = (grupo or "").strip() or "GRUPO"
    mapa = mapa or {}
    codigos = sorted({c for it in itens for c in (it.get("entidades") or [])})

    def _resolver(it):
        """Empresas do item: se veio a seleção por linha (`empresa_ids`, mesmo vazia),
        respeita ela; senão, cai no antigo esquema por código (mapa/criação)."""
        if "empresa_ids" in it:
            emps = [db.query(Empresa).filter(Empresa.id == int(i)).first()
                    for i in (it.get("empresa_ids") or [])]
            return [e for e in emps if e]
        out = []
        for c in (it.get("entidades") or []):
            eid = mapa.get(c) or mapa.get(str(c))
            e = db.query(Empresa).filter(Empresa.id == int(eid)).first() if eid else None
            out.append(e or _get_or_create_empresa(db, c, grupo))
        return out

    criadas = atualizadas = 0
    usadas = set()
    for it in itens:
        nome = (it.get("nome") or "").strip()
        if not nome:
            continue
        setor = _get_or_create_setor(db, it.get("setor"))
        emps = _resolver(it)
        for e in emps:
            usadas.add(e.razao_social)

        o = db.query(Obrigacao).filter(Obrigacao.nome == nome).first()
        if o:
            atuais = {e.id for e in o.empresas}
            for e in emps:
                if e.id not in atuais:
                    o.empresas.append(e)
            atualizadas += 1
        else:
            o = Obrigacao(
                nome=nome,
                setor_id=setor.id if setor else None,
                competencia_ref=it.get("competencia_ref") or "mes_anterior",
                regra_prazo_tipo=it.get("regra_prazo_tipo") or "ultimo_dia_util",
                regra_prazo_dia=it.get("regra_prazo_dia"),
                meses_ativos=MESES_TODOS,
                passivel_multa=bool(it.get("gera_multa")),
                ativa=True,
            )
            o.empresas = emps
            db.add(o)
            criadas += 1
    db.commit()
    return {"grupo": grupo, "criadas": criadas, "atualizadas": atualizadas,
            "empresas": sorted(usadas)}


def gerar_modelo() -> bytes:
    """XLSX-modelo do layout único de importação de obrigações."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Obrigações"
    ws.append(["Descrição da tarefa", "Competência", "Vencimento", "Empresa", "CNPJ", "Gera multa", "Setor"])
    ws.append(["Apuração do ICMS", "Mês anterior", "5º dia útil", "FOS", "12.345.678/0001-90", "Sim", "Fiscal"])
    ws.append(["Conciliação bancária", "Mês anterior", "6º dia útil", "FOS", "12.345.678/0001-90", "Não", "Financeiro"])
    ws.append(["Fechamento da folha", "Mês anterior", "Último dia útil", "EDS", "98.765.432/0001-10", "Não", "DP"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
