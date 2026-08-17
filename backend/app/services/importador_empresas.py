"""Importação de empresas em lote a partir de Excel (XLSX/XLS).

Colunas aceitas (cabeçalho flexível, sem acento/maiúsculas): razão social, CNPJ,
regime tributário, grupo (de empresas). Bônus se vierem: nome fantasia, email,
telefone, segmento. Upsert por CNPJ — se já existe, atualiza; senão, cria.
"""
import io
import re
import unicodedata
from ..models import Empresa
from .validacao import cnpj_valido


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _so_digitos(s: str) -> str:
    return re.sub(r"\D", "", str(s or ""))


# cabeçalho normalizado -> campo do modelo
COLUNAS = {
    "razao social": "razao_social", "razao": "razao_social", "empresa": "razao_social", "nome": "razao_social",
    "cnpj": "cnpj",
    "regime": "regime_tributario", "regime tributario": "regime_tributario", "tributacao": "regime_tributario",
    "grupo": "grupo", "grupo de empresas": "grupo", "grupo economico": "grupo", "grupo empresarial": "grupo",
    "nome fantasia": "nome_fantasia", "fantasia": "nome_fantasia",
    "email": "email", "e-mail": "email",
    "telefone": "telefone", "fone": "telefone", "whatsapp": "telefone",
    "segmento": "segmento",
}

REGIMES = {
    "simples": "simples_nacional", "simples nacional": "simples_nacional", "sn": "simples_nacional",
    "presumido": "lucro_presumido", "lucro presumido": "lucro_presumido", "lp": "lucro_presumido",
    "real": "lucro_real", "lucro real": "lucro_real", "lr": "lucro_real",
    "mei": "mei",
    "terceiro setor": "terceiro_setor",
    "imune": "imune",
    "isento": "isento",
}

SEGMENTOS = {
    "comercio": "comercio", "servico": "servico", "servicos": "servico",
    "comercio e servico": "comercio_servico", "comercio servico": "comercio_servico",
    "industria": "industria",
    "holding": "holding",
    "imune": "imune",
    "igreja": "igreja",
}


def _mapear_regime(v: str) -> str:
    return REGIMES.get(_norm(v), "indefinido") if v else "indefinido"


def _mapear_segmento(v: str):
    return SEGMENTOS.get(_norm(v)) if v else None


def _linhas_xlsx(conteudo: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _linhas_xls(conteudo: bytes):
    import xlrd
    sh = xlrd.open_workbook(file_contents=conteudo).sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def _ler_grade(nome: str, conteudo: bytes):
    n = (nome or "").lower()
    if n.endswith(".xls"):
        return _linhas_xls(conteudo)
    return _linhas_xlsx(conteudo)


def importar(db, nome_arquivo: str, conteudo: bytes) -> dict:
    grade = _ler_grade(nome_arquivo, conteudo)
    if not grade:
        return {"resumo": {"total": 0, "criadas": 0, "atualizadas": 0, "erros": 0}, "detalhes": []}

    # cabeçalho: primeira linha não-vazia
    idx_cab = next((i for i, r in enumerate(grade) if any(c not in (None, "") for c in r)), None)
    if idx_cab is None:
        return {"resumo": {"total": 0, "criadas": 0, "atualizadas": 0, "erros": 0}, "detalhes": []}

    cabecalho = grade[idx_cab]
    mapa = {}  # índice de coluna -> campo
    for i, titulo in enumerate(cabecalho):
        campo = COLUNAS.get(_norm(titulo))
        if campo:
            mapa[i] = campo
    if "razao_social" not in mapa.values():
        return {"erro": "Não encontrei a coluna 'Razão Social' no arquivo.",
                "resumo": {"total": 0, "criadas": 0, "atualizadas": 0, "erros": 0}, "detalhes": []}

    criadas = atualizadas = erros = 0
    detalhes = []

    for linha in grade[idx_cab + 1:]:
        if not any(c not in (None, "") for c in linha):
            continue  # linha vazia
        dados = {}
        for i, campo in mapa.items():
            valor = linha[i] if i < len(linha) else None
            if valor in (None, ""):
                continue
            if campo == "cnpj":
                dados["cnpj"] = _so_digitos(valor)
            elif campo == "regime_tributario":
                dados["regime_tributario"] = _mapear_regime(valor)
            elif campo == "segmento":
                seg = _mapear_segmento(valor)
                if seg:
                    dados["segmento"] = seg
            else:
                dados[campo] = str(valor).strip()

        razao = dados.get("razao_social")
        if not razao:
            erros += 1
            detalhes.append({"linha": razao or "(sem razão social)", "status": "erro",
                             "detalhe": "Sem razão social"})
            continue

        cnpj = dados.get("cnpj")
        if cnpj and not cnpj_valido(cnpj):
            detalhes.append({"linha": razao, "status": "aviso",
                             "detalhe": f"CNPJ inválido ({cnpj}), importado sem CNPJ."})
            dados.pop("cnpj", None)
            cnpj = None
        existente = None
        if cnpj:
            existente = next((e for e in db.query(Empresa).filter(Empresa.cnpj != None).all()
                              if _so_digitos(e.cnpj) == cnpj), None)

        if existente:
            for k, v in dados.items():
                setattr(existente, k, v)
            atualizadas += 1
            detalhes.append({"linha": razao, "status": "atualizada",
                             "detalhe": f"CNPJ {cnpj} já existia"})
        else:
            db.add(Empresa(**dados))
            criadas += 1
            detalhes.append({"linha": razao, "status": "criada", "detalhe": None})

    db.commit()
    return {"resumo": {"total": criadas + atualizadas + erros,
                       "criadas": criadas, "atualizadas": atualizadas, "erros": erros},
            "detalhes": detalhes}


def gerar_modelo() -> bytes:
    """Gera um XLSX-modelo com os cabeçalhos e uma linha de exemplo."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empresas"
    ws.append(["Razão Social", "CNPJ", "Regime Tributário", "Grupo de Empresas", "Segmento"])
    ws.append(["MARKBUILDING CONSTRUCOES LTDA", "12.345.678/0001-90", "Lucro Real", "Markbuilding", "Serviço"])
    ws.append(["GNILEB PARTICIPACOES SA", "98.765.432/0001-10", "Lucro Presumido", "Markbuilding", "Serviço"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
