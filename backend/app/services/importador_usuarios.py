"""Importação de USUÁRIOS em lote a partir de Excel (XLSX/XLS).

Colunas aceitas (cabeçalho flexível, sem acento/maiúsculas):
  nome*, email*, senha, cargo, telefone, nível (grupo), tipo, empresa, gestor.
* obrigatórias. Upsert por e-mail — se já existe, atualiza (NÃO troca a senha,
a menos que a coluna 'senha' venha preenchida). Sem 'senha' em usuário novo,
gera uma SENHA TEMPORÁRIA aleatória e a devolve no resultado (para o admin
repassar; o usuário troca depois).

Empresa (para tipo 'cliente') é resolvida por CNPJ ou razão social.
Gestor é resolvido por e-mail ou nome, em 2ª passada (pode estar em qualquer
linha do arquivo).
"""
import io
import re
import secrets
import unicodedata
from ..models import Usuario, Empresa, Setor, Grupo
from ..auth import get_password_hash


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


def _so_digitos(s: str) -> str:
    return re.sub(r"\D", "", str(s or ""))


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# cabeçalho normalizado -> campo lógico
COLUNAS = {
    "nome": "nome", "nome completo": "nome",
    "email": "email", "e-mail": "email", "e mail": "email",
    "senha": "senha", "password": "senha",
    "cargo": "cargo", "funcao": "cargo",
    "telefone": "telefone", "fone": "telefone", "celular": "telefone", "whatsapp": "telefone",
    "nivel": "grupo", "grupo": "grupo", "papel": "grupo", "perfil": "grupo", "acesso": "grupo",
    "tipo": "tipo",
    "setor": "setor", "departamento": "setor", "depto": "setor", "area": "setor",
    "empresa": "empresa", "cliente": "empresa", "empresa (cliente)": "empresa",
    "gestor": "gestor", "supervisor": "gestor", "responsavel": "gestor",
}

# rótulos amigáveis -> código do grupo (papel)
GRUPOS = {
    "admin": "admin", "administrador": "admin", "administradora": "admin",
    "gestor": "gestor", "gerente": "gestor", "coordenador": "gestor",
    "analista": "analista",
    "consulta": "consulta", "consultor": "consulta", "leitura": "consulta", "somente leitura": "consulta",
    "usuario": "usuario", "colaborador": "usuario", "operador": "usuario", "padrao": "usuario",
}

TIPOS = {
    "colaborador": "colaborador", "interno": "colaborador", "equipe": "colaborador", "funcionario": "colaborador",
    "cliente": "cliente", "externo": "cliente",
}


def _carregar_grupos(db) -> dict:
    """{nome_normalizado: slug} dos grupos ativos do banco — por slug e por label."""
    mapa = {}
    for g in db.query(Grupo).filter(Grupo.ativo == True).all():
        mapa[_norm(g.slug)] = g.slug
        mapa[_norm(g.label)] = g.slug
    return mapa


def _mapear_grupo(v: str, grupos_db: dict = None) -> str:
    """Resolve o nível: banco (slug/label) tem prioridade; depois sinônimos fixos."""
    if not v:
        return "usuario"
    n = _norm(v)
    if grupos_db and n in grupos_db:
        return grupos_db[n]
    return GRUPOS.get(n, "usuario")


def _mapear_tipo(v: str) -> str:
    return TIPOS.get(_norm(v), "colaborador") if v else "colaborador"


def _senha_temporaria() -> str:
    """Senha temporária legível (8 chars, sem caracteres ambíguos)."""
    alfabeto = "ABCDEFGHJKLMNPQRSTUVWXYZabcdefghijkmnpqrstuvwxyz23456789"
    return "".join(secrets.choice(alfabeto) for _ in range(8))


def _linhas_xlsx(conteudo: bytes):
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _linhas_xls(conteudo: bytes):
    import xlrd
    sh = xlrd.open_workbook(file_contents=conteudo).sheet_by_index(0)
    return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]


def _ler_grade(nome: str, conteudo: bytes):
    n = (nome or "").lower()
    if n.endswith(".xls"):
        return _linhas_xls(conteudo)
    return _linhas_xlsx(conteudo)


def _resolver_empresa(db, valor: str):
    """Empresa por CNPJ (dígitos) ou por razão social (normalizada)."""
    if not valor:
        return None
    dig = _so_digitos(valor)
    if len(dig) == 14:
        return next((e for e in db.query(Empresa).filter(Empresa.cnpj != None).all()
                     if _so_digitos(e.cnpj) == dig), None)
    alvo = _norm(valor)
    return next((e for e in db.query(Empresa).all() if _norm(e.razao_social) == alvo), None)


def _resolver_setor(db, valor: str):
    """Setor por nome (normalizado). Cria se não existir (import tolerante)."""
    if not valor:
        return None
    alvo = _norm(valor)
    s = next((x for x in db.query(Setor).all() if _norm(x.nome) == alvo), None)
    if not s:
        s = Setor(nome=str(valor).strip())
        db.add(s)
        db.flush()
    return s


def importar(db, nome_arquivo: str, conteudo: bytes) -> dict:
    grade = _ler_grade(nome_arquivo, conteudo)
    if not grade:
        return {"resumo": {"total": 0, "criadas": 0, "atualizadas": 0, "erros": 0}, "detalhes": []}

    idx_cab = next((i for i, r in enumerate(grade) if any(c not in (None, "") for c in r)), None)
    if idx_cab is None:
        return {"resumo": {"total": 0, "criadas": 0, "atualizadas": 0, "erros": 0}, "detalhes": []}

    mapa = {}  # índice de coluna -> campo lógico
    for i, titulo in enumerate(grade[idx_cab]):
        campo = COLUNAS.get(_norm(titulo))
        if campo:
            mapa[i] = campo
    if "nome" not in mapa.values() or "email" not in mapa.values():
        return {"erro": "O arquivo precisa das colunas 'Nome' e 'E-mail'.",
                "resumo": {"total": 0, "criadas": 0, "atualizadas": 0, "erros": 0}, "detalhes": []}

    criadas = atualizadas = erros = 0
    detalhes = []
    gestor_pendente = []  # (usuario, valor_gestor) resolvidos na 2ª passada
    grupos_db = _carregar_grupos(db)

    for linha in grade[idx_cab + 1:]:
        if not any(c not in (None, "") for c in linha):
            continue
        dados = {}
        for i, campo in mapa.items():
            valor = linha[i] if i < len(linha) else None
            if valor in (None, ""):
                continue
            dados[campo] = str(valor).strip() if not isinstance(valor, str) else valor.strip()

        nome = dados.get("nome")
        email = (dados.get("email") or "").strip().lower()
        if not nome or not email:
            erros += 1
            detalhes.append({"linha": nome or email or "(vazia)", "status": "erro",
                             "detalhe": "Sem nome ou e-mail."})
            continue
        if not EMAIL_RE.match(email):
            erros += 1
            detalhes.append({"linha": nome, "status": "erro", "detalhe": f"E-mail inválido: {email}"})
            continue

        grupo = _mapear_grupo(dados.get("grupo"), grupos_db)
        tipo = _mapear_tipo(dados.get("tipo"))
        empresa = _resolver_empresa(db, dados.get("empresa")) if tipo == "cliente" else None
        setor = _resolver_setor(db, dados.get("setor"))
        cargo = dados.get("cargo")
        telefone = dados.get("telefone")
        senha_informada = dados.get("senha")

        existente = db.query(Usuario).filter(Usuario.email == email).first()
        if existente:
            existente.nome = nome
            existente.grupo = grupo
            existente.tipo = tipo
            existente.empresa_id = empresa.id if empresa else None
            if setor:
                existente.setor_id = setor.id
            if cargo is not None:
                existente.cargo = cargo
            if telefone is not None:
                existente.telefone = telefone
            if senha_informada:
                existente.senha_hash = get_password_hash(senha_informada)
            atualizadas += 1
            detalhes.append({"linha": nome, "status": "atualizada", "detalhe": f"{email} já existia"})
            alvo = existente
        else:
            senha = senha_informada or _senha_temporaria()
            novo = Usuario(
                nome=nome, email=email, senha_hash=get_password_hash(senha),
                cargo=cargo, telefone=telefone, grupo=grupo, tipo=tipo,
                empresa_id=empresa.id if empresa else None,
                setor_id=setor.id if setor else None,
            )
            db.add(novo)
            db.flush()
            criadas += 1
            det = None if senha_informada else f"senha temporária: {senha}"
            detalhes.append({"linha": nome, "status": "criada", "detalhe": det})
            alvo = novo

        if dados.get("gestor"):
            gestor_pendente.append((alvo, dados["gestor"]))

    # 2ª passada: resolve gestor por e-mail ou nome
    for usuario, valor in gestor_pendente:
        g = None
        val = valor.strip()
        if EMAIL_RE.match(val.lower()):
            g = db.query(Usuario).filter(Usuario.email == val.lower()).first()
        if not g:
            alvo = _norm(val)
            g = next((u for u in db.query(Usuario).all() if _norm(u.nome) == alvo), None)
        if g and g.id != usuario.id:
            usuario.gestor_id = g.id
        elif not g:
            detalhes.append({"linha": usuario.nome, "status": "aviso",
                             "detalhe": f"Gestor não encontrado: {valor}"})

    db.commit()
    return {"resumo": {"total": criadas + atualizadas + erros,
                       "criadas": criadas, "atualizadas": atualizadas, "erros": erros},
            "detalhes": detalhes}


def gerar_modelo() -> bytes:
    """XLSX-modelo com cabeçalhos e exemplos."""
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuários"
    ws.append(["Nome", "E-mail", "Senha", "Cargo", "Telefone", "Nível", "Tipo", "Setor", "Empresa", "Gestor"])
    ws.append(["Ana Souza", "ana@bps4.com.br", "", "Analista Fiscal", "(21) 99999-0000",
               "Analista", "Colaborador", "Fiscal", "", "carlos@bps4.com.br"])
    ws.append(["Carlos Lima", "carlos@bps4.com.br", "", "Gestor Contábil", "(21) 98888-0000",
               "Gestor", "Colaborador", "Contábil", "", ""])
    ws.append(["João Cliente", "joao@empresa.com.br", "", "", "",
               "Consulta", "Cliente", "", "12.345.678/0001-90", ""])
    for col, w in zip("ABCDEFGHIJ", (22, 28, 14, 20, 18, 12, 14, 16, 26, 26)):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
